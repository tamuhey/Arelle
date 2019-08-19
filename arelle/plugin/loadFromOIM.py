'''
loadFromExcel.py is an example of a plug-in that will load an extension taxonomy from Excel
input and optionally save an (extension) DTS.

(c) Copyright 2016 Mark V Systems Limited, All rights reserved.

Example to run from web server:

1) POSTing excel in a zip, getting instance and log back in zip:

   curl -k -v -X POST "-HContent-type: application/zip" -T /Users/hermf/Documents/blahblah.xlsx.zip "localhost:8080/rest/xbrl/open?media=zip&file=WIP_DETAILED_3.xlsx&plugins=loadFromOIM&saveOIMinstance=myinstance.xbrl" -o out.zip
   
2) POSTing json within an archive of XII test cases and log and instance back in zip

   curl -k -v -X POST "-HContent-type: application/zip" -T test.zip  "localhost:8080/rest/xbrl/open?media=zip&file=100-json/helloWorld.json&plugins=loadFromOIM&saveOIMinstance=myinstance.xbrl" -o out.zip


'''
import os, sys, io, time, re, traceback, json, csv, logging, math, zipfile, datetime, isodate
from lxml import etree
from collections import defaultdict, OrderedDict
from arelle.ModelDocument import Type, create as createModelDocument
from arelle import XbrlConst, ModelDocument, ModelXbrl, ValidateXbrlDimensions
from arelle.ModelDocument import Type, create as createModelDocument
from arelle.ModelValue import qname, dateTime, DATETIME, yearMonthDuration, dayTimeDuration
from arelle.PrototypeInstanceObject import DimValuePrototype
from arelle.PythonUtil import attrdict
from arelle.UrlUtil import isHttpUrl
from arelle.XbrlConst import (qnLinkLabel, standardLabelRoles, qnLinkReference, standardReferenceRoles,
                              qnLinkPart, gen, link, defaultLinkRole,
                              conceptLabel, elementLabel, conceptReference
                              )
from arelle.XmlUtil import addChild, addQnameValue, copyIxFootnoteHtml, setXmlns
from arelle.XmlValidate import NCNamePattern, validate as xmlValidate

nsOim = "http://www.xbrl.org/CR/2018-12-12"
nsOims = (nsOim,
          "http://www.xbrl.org/WGWD/YYYY-MM-DD",
          "http://www.xbrl.org/PWD/2016-01-13/oim",
          "http://www.xbrl.org/WGWD/YYYY-MM-DD/oim"
         )
nsOimCes = (
        "http://www.xbrl.org/WGWD/YYYY-MM-DD/oim-common/error",
    )
jsonDocumentTypes = (
        "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-json",
    )
csvDocumentTypes = (
        "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-csv",
        "http://xbrl.org/WGWD/YYYY-MM-DD/xbrl-csv",
    )
         


XLINKTYPE = "{http://www.w3.org/1999/xlink}type"
XLINKLABEL = "{http://www.w3.org/1999/xlink}label"
XLINKARCROLE = "{http://www.w3.org/1999/xlink}arcrole"
XLINKFROM = "{http://www.w3.org/1999/xlink}from"
XLINKTO = "{http://www.w3.org/1999/xlink}to"
XLINKHREF = "{http://www.w3.org/1999/xlink}href"
XMLLANG = "{http://www.w3.org/XML/1998/namespace}lang"

JSONdocumentType = "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-json"

OIMReservedAliasURIs = {
    "xbrl": nsOims,
    "xsd": (XbrlConst.xsd,),
    "enum2": XbrlConst.enum2s,
    "oimce": nsOimCes,
    "xbrli": (XbrlConst.xbrli,),
    "xsd": (XbrlConst.xsd,),
    "xbrlje": [ns + "/xbrl-json/error" for ns in nsOims]
    }
OIMReservedURIAlias = dict(
    (uri, alias)
    for alias, uris in OIMReservedAliasURIs.items()
    for uri in uris)

ENTITY_NA_QNAME = qname("https://xbrl.org/entities", "NA")
EMPTY_DICT = {}

DUPJSONKEY = "!@%duplicateKeys%@!"
DUPJSONVALUE = "!@%duplicateValues%@!"

EMPTYDICT = {}

PrefixedQName = re.compile(
                 "[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
                  r"[_\-\." 
                  "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*:"
                 "[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
                  r"[_\-\." 
                  "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*")
UnitPrefixedQNameSubstitutionChar = "\x07" # replaces PrefixedQName in unit pattern
UnitPattern = re.compile(
                # QNames are replaced by \x07 in these expressions
                # numerator only (no parentheses)
                "(^\x07$)|(^\x07([*]\x07)+$)|"
                # numerator and optional denominator, with parentheses if more than one term in either
                "(^((\x07)|([(]\x07([*]\x07)+[)]))([/]((\x07)|([(]\x07([*]\x07)+[)])))?$)"
                )

xlUnicodePattern = re.compile("_x([0-9A-F]{4})_")

def xlUnicodeChar(match):
    return chr(int(match.group(1), 16))
    
def xlValue(cell): # excel values may have encoded unicode, such as _0000D_
    v = cell.value
    if isinstance(v, str):
        return xlUnicodePattern.sub(xlUnicodeChar, v).replace('\r\n','\n').replace('\r','\n')
    return v

class OIMException(Exception):
    def __init__(self, code, message, **kwargs):
        self.code = code
        self.message = message
        self.msgArgs = kwargs
        self.args = ( self.__repr__(), )
    def __repr__(self):
        return _('[{0}] exception {1}').format(self.code, self.message % self.msgArgs)

class NotOIMException(Exception):
    def __init__(self, **kwargs):
        self.args = ( self.__repr__(), )
    def __repr__(self):
        return _('[NotOIM] not an OIM document')

def csvCellValue(cellValue):
    if cellValue == "#nil":
        return None
    elif cellValue == "#empty":
        return ""
    elif isinstance(cellValue, str) and cellValue.startswith("##"):
        return cellValue[1:]
    else:
        return cellValue

PER_ISO = 0
PER_INCLUSIVE_DATES = 1
PER_SINGLE_DAY = 2
PER_MONTH = 3
PER_YEAR = 4
PER_QTR = 5
PER_HALF = 6
PER_WEEK = 7
ONE_DAY = dayTimeDuration("P1D")
ONE_MONTH = yearMonthDuration("P1M")
ONE_YEAR = yearMonthDuration("P1Y")
ONE_QTR = yearMonthDuration("P3M")
ONE_HALF = yearMonthDuration("P6M")
    
periodForms = ((PER_ISO, re.compile("([0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{2}:[0-9]{2}:[0-9]{2}(/[0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{2}:[0-9]{2}:[0-9]{2})?)$")),
               (PER_INCLUSIVE_DATES, re.compile("([0-9]{4}-[0-9]{2}-[0-9]{2})[.][.]([0-9]{4}-[0-9]{2}-[0-9]{2})$")),
               (PER_SINGLE_DAY, re.compile("([0-9]{4})-([0-9]{2})-([0-9]{2})$")),
               (PER_MONTH,  re.compile("([0-9]{4})-([0-9]{2})$")),
               (PER_YEAR, re.compile("([0-9]{4})$")),
               (PER_QTR, re.compile("([0-9]{4})Q([1-4])$")),
               (PER_HALF, re.compile("([0-9]{4})H([1-2])$")),
               (PER_WEEK, re.compile("([0-9]{4}W[1-5]?[0-9])$")))
def csvPeriod(cellValue, startOrEnd):
    if cellValue in ("", "#none"):
        return None
    isoDuration = None
    for perType, perFormMatch in periodForms:
        m = perFormMatch.match(cellValue)
        if m:
            try:
                if perType == PER_ISO:
                    return cellValue
                elif perType == PER_INCLUSIVE_DATES:
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(dateTime(m.group(1)), dateTime(m.group(2)) + ONE_DAY)
                elif perType == PER_SINGLE_DAY:
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(dateTime(m.group(1)), dateTime(m.group(1)) + ONE_DAY)
                elif perType == PER_MONTH:
                    moStart = dateTime(m.group(1) + "-00")
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(moStart, moStart + ONE_MONTH)
                elif perType == PER_YEAR:
                    yrStart = dateTime(m.group(1) + "-00-00")
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(yrStart, yrStart + ONE_YEAR)
                elif perType == PER_QTR:
                    qtrStart = dateTime(m.group(1) + "-{:02}-00".format(int(m.group(2))*3 - 3))
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(qtrStart, qtrStart + ONE_QTR)
                elif perType == PER_HALF:
                    qtrStart = dateTime(m.group(1) + "-{:02}-00".format(int(m.group(2))*6 - 6))
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(qtrStart, qtrStart + ONE_HALF)
                elif perType == PER_WEEK:
                    weekStart = dateTime(isodate.parse_date(m.group(1)))
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(weekStart, qtrStart + datetime.timedelta(7))
            except ValueError:
                return None
        if isoDuration:
            if startOrEnd == "start":
                return isoDuration.partition("/")[0]
            elif startOrEnd == "end":
                return isoDuration.partition("/")[2]
            return isoDuration
    return None
    

def loadFromOIM(cntlr, error, warning, modelXbrl, oimFile, mappedUri, oimObject=None):
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
    from arelle import ModelDocument, ModelXbrl, XmlUtil
    from arelle.ModelDocument import ModelDocumentReference
    from arelle.ModelValue import qname
    
    _return = None # modelDocument or an exception
    
    try:
        currentAction = "initializing"
        startingErrorCount = len(modelXbrl.errors) if modelXbrl else 0
        startedAt = time.time()
        
        if not modelXbrl.fileSource.isArchive and os.path.isabs(oimFile):
            # allow relative filenames to loading directory
            priorCWD = os.getcwd()
            os.chdir(os.path.dirname(oimFile))
        else:
            priorCWD = None
            
        currentAction = "determining file type"
        isXL = oimFile.endswith(".xlsx") or oimFile.endswith(".xls")
        isJSON = False
        isCSV = False # oimFile.endswith(".csv") # this option is not currently supported
        isCSVorXL = isCSV or isXL
        instanceFileName = os.path.splitext(oimFile)[0] + ".xbrl"
        _csvwContext = None
        anonymousFactId = 0 
        
        if not isXL: # try as JSON
            currentAction = "loading and parsing JSON OIM file"
            def loadDict(keyValuePairs):
                _dict = OrderedDict() # preserve fact order in resulting instance
                _valueKeyDict = {}
                for key, value in keyValuePairs:
                    if isinstance(value, dict):
                        if DUPJSONKEY in value:
                            for _errKey, _errValue, _otherValue in value[DUPJSONKEY]:
                                if key in ("namespaces", "linkTypes", "groups"):
                                    error("oimce:multipleURIsForAlias",
                                                    _("The %(map)s alias %(prefix)s is used on uri %(uri1)s and uri %(uri2)s"),
                                                    modelObject=modelXbrl, map=key, prefix=_errKey, uri1=_errValue, uri2=_otherValue)
                                else:
                                    error("oim:duplicateKeys",
                                                    _("The %(obj)s key %(key)s is used on multiple objects"),
                                                    modelObject=modelXbrl, obj=key, key=_errKey)
                            del value[DUPJSONKEY]
                        if DUPJSONVALUE in value:
                            for _errValue, _errKey, _otherKey in value[DUPJSONVALUE]:
                                if key in ("namespaces", "linkTypes", "groups"):
                                    error("oimce:multipleAliasesForURI",
                                                    _("The %(map)s value %(uri)s is used on alias %(alias1)s and alias %(alias2)s"),
                                                    modelObject=modelXbrl, map=key, uri=_errValue, alias1=_errKey, alias2=_otherKey)
                            del value[DUPJSONVALUE]
                        if key in ("namespaces", "linkTypes", "groups"):
                            for _key, _value in value.items():
                                if _key == "":
                                    error("oimce:invalidEmptyURIAlias",
                                                    _("The %(map)s empty string MUST NOT be used as a URI alias: uri %(uri)s"),
                                                    modelObject=modelXbrl, map=key, prefix="", uri=_value)
                                elif key == "namespaces" and _key in OIMReservedAliasURIs and _value not in OIMReservedAliasURIs[_key]:
                                    error("oimce:invalidURIForReservedAlias",
                                                    _("The namespaces URI %(uri)s is used on standard alias %(alias)s which requires URI %(standardUri)s"),
                                                    modelObject=modelXbrl, alias=_key, uri=_value, standardUri=OIMReservedAliaseURIs[_key][0])
                                elif key == "namespaces" and _value in OIMReservedURIAlias and _key != OIMReservedURIAlias[_value]:
                                    error("oimce:invalidPrefixForURIWithReservedAlias",
                                                    _("The namespaces URI %(uri)s is bound to alias %(key)s instead of standard alias %(alias)s"),
                                                    modelObject=modelXbrl, key=_key, uri=_value, alias=OIMReservedURIAlias[_value])
                    if key in _dict:
                        if DUPJSONKEY not in _dict:
                            _dict[DUPJSONKEY] = []
                        _dict[DUPJSONKEY].append((key, value, _dict[key]))
                    elif isinstance(value, str) and value in _valueKeyDict:
                        if DUPJSONVALUE not in _dict:
                            _dict[DUPJSONVALUE] = []
                        _dict[DUPJSONVALUE].append((value, key, _valueKeyDict[value]))
                    else:
                        _dict[key] = value
                        if isinstance(value, str):
                            _valueKeyDict[value] = key
                return _dict
            if oimObject is None:
                _file, encoding = modelXbrl.fileSource.file(oimFile, encoding='utf-8')
                with _file as f:
                    oimObject = json.load(f, object_pairs_hook=loadDict)
            documentInfo = oimObject.get("documentInfo", {})
            documentType = documentInfo.get("documentType")
            if documentType in jsonDocumentTypes:
                isJSON = True
                errPrefix = "xbrlje"
            elif documentType in csvDocumentTypes:
                isCSV = True
                errPrefix = "xbrlce"
            else:
                raise OIMException("arelle:unrecognizedDocumentType", 
                                   _("OIM document type is not recognized: %(documentType)s"),
                                   documentType=documentType)
            
            missing = [t 
                       for t in (("documentInfo", "facts") if isJSON else ("documentInfo", "tables") )
                       if t not in oimObject]
            missing += [t 
                        for t in (("documentType", "namespaces", "features") if isJSON else ("documentType", "namespaces", "taxonomy")) 
                        if t not in documentInfo]
            if isCSV:
                missing += ["{}:{}".format(n,t) 
                            for n,tbl in oimObject.get("tables",{}).items()
                            for t in ("url", "factColumns", "tableDimensions") 
                            if t not in tbl]
            if missing:
                raise OIMException("{}:invalidMeadataStructure".format(errPrefix), 
                                   _("Required element(s) are missing from metadata: %(missing)s"),
                                   missing = ", ".join(missing))
            oimDocumentInfo = oimObject["documentInfo"]
            currentAction = "identifying Metadata objects"
            taxonomyRefs = oimDocumentInfo["taxonomy"]
            namespaces = oimDocumentInfo["namespaces"]
            linkTypes = oimDocumentInfo.get("linkTypes", EMPTY_DICT)
            groups = oimDocumentInfo.get("groups",EMPTY_DICT)
            if isJSON:
                featuresDict = oimDocumentInfo["features"]
                factItems = oimObject["facts"].items()
                footnotes = oimObject["facts"].values() # shares this object
            else: # isCSV
                reportDimensions = documentInfo.get("reportDimensions", {})
                reportDecimals = documentInfo.get("decimals", None)
                tables = oimObject["tables"]
                footnotes = oimObject.get("links", {})
                if sys.version[0] >= '3':
                    csvOpenMode = 'w'
                    csvOpenNewline = ''
                else:
                    csvOpenMode = 'wb' # for 2.7
                    csvOpenNewline = None

        if isCSV:
            currentAction = "loading CSV facts tables"
            _dir = os.path.dirname(oimFile)
            footnoteRefFactIds = defaultdict(set)
            anonymousFootnoteId = 0 # integer always for anonymous (same row) footnotes

            def csvFacts():
                for tableId, table in tables.items():
                    tableDimensions = table.get("tableDimensions", {})
                    tableDecimals = table.get("decimals", None)
                    tableUrl = table["url"]
                    # compile column dependencies
                    factDimensions = {}
                    factDecimals = {}
                    for colId, colProperties in table["factColumns"].items():
                        factDimensions[colId] = colProperties.get("columnDimensions", {})
                        factDecimals[colId] = colProperties.get("decimals", None)
                    _file, encoding = modelXbrl.fileSource.file(os.path.join(_dir, tableUrl), encoding='utf-8-sig')
                    with _file as f:
                        csvReader = csv.reader(f)
                        for rowIndex, row in enumerate(csvReader):
                            if rowIndex == 0:
                                header = row
                                colNameIndex = dict((name, colIndex) for colIndex, name in enumerate(row))
                                idColIndex = colNameIndex.get(table.get("rowIdColumn"))
                            else:
                                for colIndex, colValue in enumerate(row):
                                    if colIndex >= len(header):
                                        continue
                                    colName = header[colIndex]
                                    if colName not in factDimensions:
                                        continue # not a fact column
                                    id = None
                                    fact = {}
                                    # if this is an id column
                                    if colValue not in (None, ""):
                                        fact["value"] = csvCellValue(colValue)
                                        fact["dimensions"] = colFactDims = factDimensions[colName].copy()
                                        for inheritedDims in tableDimensions, reportDimensions:
                                            for dimName, dimValue in inheritedDims.items():
                                                if dimName not in colFactDims:
                                                    colFactDims[dimName] = inheritedDims[dimName]
                                        # resolve column-relative dimensions
                                        dimsNamesToRemove = []
                                        for dimName, dimValue in colFactDims.items():
                                            if isinstance(dimValue, str) and dimValue.startswith("$"):
                                                dimValue = dimValue[1:]
                                                if not dimValue.startswith("$"):
                                                    dimValue, _sep, dimAttr = dimValue.partition("@")
                                                    if dimValue in colNameIndex:
                                                        dimValue = row[colNameIndex[dimValue]]
                                                    if dimName == "period":
                                                        dimValue = csvPeriod(dimValue, dimAttr)
                                                if dimValue in ("", "#none"):
                                                    dimsNamesToRemove.append(dimName)
                                                else:
                                                    colFactDims[dimName] = dimValue
                                            elif dimValue == "#none":
                                                dimsNamesToRemove.append(dimName)
                                        for dimName in dimsNamesToRemove:
                                            del colFactDims[dimName]
                                        if colName in factDecimals:
                                            fact["decimals"] = factDecimals[colName]
                                        elif tableDecimals is not None:
                                            fact["decimals"] = tableDecimals
                                        else:
                                            fact["decimals"] = reportDecimals
                                        id = "{}.{}.{}".format(
                                            tableId, 
                                            row[idColIndex] if idColIndex is not None and idColIndex < len(row) else rowIndex, 
                                            colName)
                                        yield (id, fact)

            factItems = csvFacts()
                
        elif isXL:
            errPrefix = "xbrlce" # use same prefix as CSV since workbook is a use of xBRL-CSV specification
            currentAction = "identifying workbook input worksheets"
            _file, = modelXbrl.fileSource.file(oimFile, binary=True)
            with _file as f:
                oimWb = load_workbook(f, data_only=True)
            sheetNames = oimWb.get_sheet_names()
            if (not any(sheetName == "aliases" for sheetName in sheetNames) or
                not any(sheetName == "taxonomyRefs" for sheetName in sheetNames) or
                not any("metadata" in sheetName for sheetName in sheetNames)):
                raise OIMException("xbrlwe:missingWorkbookWorksheets", 
                                   _("Unable to identify worksheet tabs for taxonomyRefs, aliases or metadata"))
            currentAction = "loading worksheet: taxonomyRefs"
            taxonomyRefs = []
            for i, row in enumerate(oimWb["taxonomyRefs"]):
                if i == 0:
                    header = [xlValue(col) for col in row]
                elif any(col.value is not None for col in row): # skip entirely empty rows
                    taxonomyRefs.append(dict((header[j], col.value) for j, col in enumerate(row)))
            currentAction = "loading worksheet: namespaces"
            aliasesList = []
            for i, row in enumerate(oimWb["namespaces"]):
                if i == 0:
                    header = dict((xlValue(col),i) for i,col in enumerate(row))
                elif any(col.value is not None for col in row): # skip entirely empty rows
                    aliasesList.append((row[header["prefix"]].value, row[header["URI"]].value))
                    tableMetadata = OrderedDict() # list of rows per table name
            if "metadata" in sheetNames:
                currentAction = "loading worksheet: metadata"
                topLevelProperties = {}
                missingTables = set()
                missingRanges = set()
                for i, row in enumerate(oimWb["metadata"]):
                    if i == 0:
                        metaTitles = [col.value for col in row]
                        metaHdr = dict((title,i) for i,title in enumerate(metaTitles))
                        missingCols = {"table", "column name", "column type"} - set(metaHdr.keys())
                        if missingCols:
                            raise OIMException("xbrlwe:missingMetadataColumns", 
                                               _("Required columns missing: %(missing)s"),
                                               missing=", ".join(sorted(missingCols)))
                        metaColPropCols = [] # pairs of property name and applies to metadata column indices
                        for i, col in enumerate(metaTitles):
                            if col == "column property":
                                metaColPropCols.append([i, None])
                            elif col == "applies to" and metaColPropCols:
                                metaColPropCols[-1][1] = i
                    elif any(col.value is not None for col in row): # skip entirely empty rows
                        tableRangeName = xlValue(row[metaHdr["table"]])
                        if not tableRangeName: # top level properties
                            for col, iCol in metaHdr.items():
                                value = xlValue(row[iCol])
                                if col not in ("table", "column name", "column type", "column property", "applies to") and value is not None:
                                    topLevelProperties[col] = value
                        else:
                            if tableRangeName not in tableMetadata: # first encounter of tableRangeName, check if it's a range
                                table, _sep, namedRange = tableRangeName.partition('!')
                                if table not in sheetNames:
                                    missingTables.add(table)
                                elif namedRange:
                                    if namedRange in oimWb.defined_names:
                                        defn = oimWb.defined_names[namedRange]
                                        if defn.type != "RANGE":
                                            raise OIMException("xbrlwe:unusableRange", 
                                                               _("Referenced range does not refer to a range: %(tableRange)s"),
                                                               tableRange=tableRangeName)
                                        elif any(_table != table for _table, cellsRange in defn.destinations):
                                            raise OIMException("xbrlwe:unusableRange", 
                                                               _("Referenced range refers to a different table: %(tableRange)s"),
                                                               tableRange=tableRangeName)
                                    else:
                                        missingRanges.add(tableRangeName)
                            tableMetadata.setdefault(tableRangeName, []).append(row)
                if missingTables:
                    raise OIMException("xbrlwe:missingTables", 
                                       _("Referenced table tab(s): %(missing)s"),
                                       missing=", ".join(sorted(missingTables)))
                if missingRanges:
                    raise OIMException("xbrlwe:missingTableNamedRanges", 
                                       _("Referenced named ranges tab(s): %(missing)s"),
                                       missing=", ".join(sorted(missingRanges)))
                facts = OrderedDict()
                footnotes = []
                footnoteRefFactIds = defaultdict(set)
                anonymousFootnoteId = 0 # integer always for anonymous (same row) footnotes
                # process by table
                for tableRangeName, tableRows in tableMetadata.items():
                    # compile column dependencies
                    propertyCols = []
                    factCols = []
                    footnoteCols = []
                    # columns for tableName worksheet
                    tableLevelProperties = {}
                    tableLevelProperties = dict((col, xlValue(row[iCol]))
                                                for row in tableRows
                                                if row[metaHdr["column name"]].value in (None, '')
                                                for col, iCol in metaHdr.items()
                                                if col not in ("table", "column name", "datatype", "column type", "column property", "applies to") and row[iCol].value is not None)
                    colDefs = [] # column definitions
                    for iCol, row in enumerate(tableRows):
                        colDef = attrdict(colName=xlValue(row[metaHdr["column name"]]),
                                          colType=xlValue(row[metaHdr["column type"]]),
                                          colProperty={},
                                          producedProperties=set())
                        colDefs.append(colDef)
                        # find column properties that apply to lists of column names
                        _firstColProp = True
                        for colPropName, colPropAppliesTo in metaColPropCols:
                            _colProperty = xlValue(row[colPropName])
                            if _colProperty:
                                colDef.colProperty[_colProperty] = (xlValue(row[colPropAppliesTo]) or "").split if colPropAppliesTo is not None else []
                                if _firstColProp:
                                    propertyCols.append(iCol)
                                    _firstColProp = False
                            colDef.producedProperties.add(colPropName)
                        # next apply properties specified for just this column
                        for col, hCol in metaHdr.items():
                            if col not in ("table", "column name", "datatype", "column type", "column property", "applies to") and row[hCol].value is not None:
                                colDef.colProperty[col] = xlValue(row[hCol])
                        if colDef.colType in ("tupleFact", "simpleFact", "numericSimpleFact", "textSimpleFact"):
                            factCols.append(iCol)
                            if "footnoteFor" in colDef.colProperty:
                                footnoteCols.append(iCol) # in-row fact footnotes
                        elif colDef.colType in ("textFootnote", "factFootnote"):
                            footnoteCols.append(iCol)
                    tableProperties = topLevelProperties.copy()
                    for propertyName, propertyValue in tableLevelProperties.items():
                        if propertyName == "deleteInheritedProperties":
                            for prop in propertyValue:
                                tableProperties.pop(prop, None)
                    for propertyName, propertyValue in tableLevelProperties.items():
                        if propertyName in ("footnoteRefs",):
                            tableProperties[propertyName].extend(propertyValue) 
                        elif propertyName != "deleteInheritedProperties":
                            tableProperties[propertyName] = propertyValue
                    tupleIds = set()
                    tableName, _sep, namedRange = tableRangeName.partition('!')
                    rangeRows = []
                    if tableName in oimWb:
                        ws = oimWb[tableName]
                        if namedRange and oimWb.defined_names[namedRange].type == "RANGE":
                            for _tableName, cells_range in oimWb.defined_names[namedRange].destinations:
                                if _tableName == tableName:
                                    rows = ws[cells_range]
                                    if isinstance(rows, Cell):
                                        rangeRows.append((rows, ))
                                    else:
                                        rangeRows.extend(rows)
                        else:
                            rangeRows = ws
                    for i, row in enumerate(rangeRows):
                        if i == 0 and not namedRange:
                            tblHdr = dict((col.value,j) for j,col in enumerate(row))
                        elif any(col.value is not None for col in row):
                            colProperties = tableProperties.copy()
                            specificColProperties = defaultdict(dict)
                            for iCol in propertyCols:
                                value = xlValue(row[iCol])
                                for _property, _colNames in colDefs[iCol].colProperty.items():
                                    if _colNames and isinstance(_colNames, list):
                                        for _colName in _colNames:
                                            specificColProperties[_colName][_property] = value
                                    else:
                                        specificColProperties['*'][_property] = value
                            for iCol in footnoteCols:
                                cellValue = xlValue(row[iCol])
                                if cellValue is None or cellValue == "": # no fact produced for this cell
                                    continue
                                colDef = colDefs[iCol]
                                cellProperties = (colProperties, 
                                                  specificColProperties.get("*", EMPTYDICT),
                                                  specificColProperties[colDef.colName],
                                                  colDef.colProperty)              
                                footnote = {}
                                if colDef.colType == "textFootnote":
                                    footnote["footnote"] = cellValue
                                elif colDef.colType == "factFootnote":
                                    footnote["factRef"] = cellValue
                                for _properties in cellProperties:
                                    if _properties:
                                        for propertyName, propertyValue in _properties.items():
                                            if propertyName == "deleteInheritedProperties":
                                                for prop in propertyValue:
                                                    footnote({"footnoteGroup":"group"}.pop(prop,prop), None)
                                        for propertyName, propertyValue in _properties.items():
                                            if propertyName != "deleteInheritedProperties" and propertyName in ("footnoteId", "footnoteType", "footnoteGroup"):
                                                footnote[{"footnoteGroup":"group"}.get(propertyName,propertyName)] = propertyValue
                                if "footnoteId" not in footnote:
                                    anonymousFootnoteId += 1
                                    footnote["footnoteId"] = "_f_{:02}".format(anonymousFootnoteId)
                                if colDef.colType.endswith("Fact"):
                                    anonymousFootnoteId += 1
                                    footnote["factRef"] = "_f_{:02}".format(anonymousFootnoteId)
                                    refs = specificColProperties[colDef.get("name")].setdefault("footnoteRefs", [])
                                    refs.append(footnote["factRef"])
                                for footnoteForCol in (colDef.colProperty.get("footnoteFor") or "").split():
                                    refs = specificColProperties[footnoteForCol].setdefault("footnoteRefs", [])
                                    refs.append(footnote["footnoteId"])                                        
                                footnotes.append(footnote)
                            for iCol in factCols:
                                if iCol >= len(row):
                                    continue
                                cellValue = xlValue(row[iCol])
                                if cellValue is None or cellValue == "": # no fact produced for this cell
                                    continue
                                colDef = colDefs[iCol]
                                cellProperties = (colProperties, 
                                                  specificColProperties.get("*", EMPTYDICT),
                                                  specificColProperties.get(colDef.colName, EMPTYDICT),
                                                  colDef.colProperty)
                                fact = {"dimensions": {}}
                                inapplicableProperties = set()
                                if colDef.colType == "tupleFact":
                                    if cellValue:
                                        if cellValue in tupleIds:
                                            continue # don't duplicate parent tuple
                                        fact["id"] = cellValue
                                        tupleIds.add(cellValue) # prevent tuple duplication
                                elif colDef.colType in ("simpleFact", "numericSimpleFact", "textSimpleFact"):
                                    fact["value"] = csvCellValue(cellValue)
                                        
                                if colDef.colType == "tupleFact":
                                    inapplicableProperties.update(oimSimpleFactProperties)
                                    for _properties in cellProperties:
                                        for propertyName, propertyValue in _properties.items():
                                            if not propertyName.startswith(oimPrefix) and propertyName != "deleteInheritedProperties":
                                                inapplicableProperties.add(propertyName)
                                        
                                # block any row property produced by this column from this column's fact
                                inapplicableProperties.update(colDef.producedProperties)
                                        
                                footnoteRefs = set()
                                for _properties in cellProperties:
                                    if _properties:
                                        for propertyName, propertyValue in _properties.items():
                                            if propertyName == "deleteInheritedProperties":
                                                for prop in propertyValue:
                                                    if ":" in prop:
                                                        fact["dimensions"].pop(prop, None)
                                                    elif prop == "footnoteRefs":
                                                        footnoteRefs.clear()
                                                    elif prop not in ("datatype",):
                                                        fact.pop(prop, None)
                                        for propertyName, propertyValue in _properties.items():
                                            if propertyName != "deleteInheritedProperties" and propertyName not in inapplicableProperties and propertyValue  != "":
                                                if ":" in propertyName:
                                                    fact["dimensions"][propertyName] = csvCellValue(propertyValue)
                                                elif propertyName == "footnoteRefs":
                                                    if isinstance(propertyValue, str): # obtained from column of blank-separated refs
                                                        propertyValue = propertyValue.split()
                                                    footnoteRefs.update(propertyValue)
                                                elif propertyName not in ("datatype",):
                                                    fact[propertyName] = propertyValue
                                if "id" not in fact:
                                    anonymousFactId += 1
                                    fact["id"] = "_f_{:02}".format(anonymousFactId)
                                factId = fact["id"]
                                if footnoteRefs:
                                    for footnoteRef in footnoteRefs:
                                        footnoteRefFactIds[footnoteRef].add(factId)
                                facts[fact["id"]] = fact
                del tupleIds
                            
    
        currentAction = "identifying default dimensions"
        if modelXbrl is not None:
            ValidateXbrlDimensions.loadDimensionDefaults(modelXbrl) # needs dimension defaults 
        
        currentAction = "validating OIM"
        
        # create the instance document
        currentAction = "creating instance document"
        if modelXbrl: # pull loader implementation
            modelXbrl.blockDpmDBrecursion = True
            modelXbrl.modelDocument = _return = createModelDocument(
                  modelXbrl, 
                  Type.INSTANCE,
                  instanceFileName,
                  schemaRefs=taxonomyRefs,
                  isEntry=True,
                  initialComment="extracted from OIM {}".format(mappedUri),
                  documentEncoding="utf-8")
            modelXbrl.modelDocument.inDTS = True
        else: # API implementation
            modelXbrl = ModelXbrl.create(
                cntlr.modelManager, 
                Type.INSTANCE, 
                instanceFileName, 
                schemaRefs=taxonomyRefs, 
                isEntry=True, 
                initialComment="extracted from OIM {}".format(mappedUri))
            _return = modelXbrl.modelDocument
        
        firstCntxUnitFactElt = None
            
        cntxTbl = {}
        unitTbl = {}
        xbrlNoteTbl = {} # fact ID: note fact
            
        currentAction = "creating facts"
        factNum = 0 # for synthetic fact number
        if not isCSV:
            syntheticFactFormat = "_f{{:0{}}}".format(int(math.log10(len(facts)))) #want 
        else:
            syntheticFactFormat = "_f{}" #want 
        
        for id, fact in factItems:
            
            dimensions = fact.get("dimensions", EMPTYDICT)
            if "concept" not in dimensions:
                error("{}:conceptQName".format(errPrefix),
                                _("The concept QName could not be determined, dimension \"concept\" is missing."),
                                modelObject=modelXbrl)
                return
            conceptSQName = dimensions["concept"]
            conceptPrefix = conceptSQName.rpartition(":")[0]
            if conceptSQName == "xbrl:note":
                xbrlNoteTbl[id] = fact
                continue
            elif not NCNamePattern.match(conceptPrefix):
                error("oimce:invalidSQNamePrefix",
                                _("The prefix of %(concept)s must match the NCName lexical pattern"),
                                modelObject=modelXbrl, concept=conceptSQName)
                continue
            elif conceptPrefix not in namespaces:
                error("oimce:unboundSQNamePrefix",
                      _("The concept SQName prefix was not defined in namespaces: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptSQName)
                continue
            conceptQn = qname(conceptSQName, namespaces)
            concept = modelXbrl.qnameConcepts.get(conceptQn)
            if concept is None:
                error("xbrl:schemaImportMissing",
                      _("The concept QName could not be resolved with available DTS: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptQn)
                return
            attrs = {}
            if concept.isItem:
                missingDimensions = []
                if "entity" not in dimensions: 
                    missingDimensions.append(oimEntity)
                if "xbrl:start" in dimensions and "xbrl:end" in dimensions:
                    pass
                if missingDimensions:
                    error("{}:missingDimensions".format(errPrefix),
                                    _("The concept %(element)s is missing dimensions %(missingDimensions)s"),
                                    modelObject=modelXbrl, element=conceptQn, missingDimensions=", ".join(missingDimensions))
                    return
                if "language" in dimensions:
                    attrs["{http://www.w3.org/XML/1998/namespace}lang"] = dimensions["language"]
                if "entity" in dimensions:
                    entityAsQn = qname(dimensions["entity"], namespaces) or ENTITY_NA_QNAME
                else:
                    entityAsQn = ENTITY_NA_QNAME
                if "xbrl:start" in dimensions and "xbrl:end" in dimensions:
                    # CSV/XL format
                    period = dimensions["xbrl:start"]
                    if period != dimensions["xbrl:end"]:
                        period += "/" + dimensions["xbrl:end"]
                elif "period" in dimensions:
                    period = dimensions["period"]
                    if period is None:
                        period = "forever"
                    elif not re.match(r"\d{4,}-[0-1][0-9]-[0-3][0-9]T([0-1][0-9]:[0-5][0-9]:[0-5][0-9])"
                                      r"(/\d{4,}-[0-1][0-9]-[0-3][0-9]T([0-1][0-9]:[0-5][0-9]:[0-5][0-9]))?", period):
                        error("{}:periodDateTime".format(errPrefix),
                              _("The concept %(element)s has a lexically invalid period dateTime %(periodError)s"),
                              modelObject=modelXbrl, element=conceptQn, periodError=periodDate)
                else:
                    period = "forever"
                cntxKey = ( # hashable context key
                    ("periodType", concept.periodType),
                    ("entity", entityAsQn),
                    ("period", period)) + tuple(sorted(
                        (dimName, dimVal["value"] if isinstance(dimVal,dict) else dimVal) 
                        for dimName, dimVal in dimensions.items()
                        if ":" in dimName))
                if cntxKey in cntxTbl:
                    _cntx = cntxTbl[cntxKey]
                else:
                    cntxId = 'c-{:02}'.format(len(cntxTbl) + 1)
                    qnameDims = {}
                    for dimName, dimVal in dimensions.items():
                        if ":" in dimName:
                            dimQname = qname(dimName, namespaces)
                            dimConcept = modelXbrl.qnameConcepts.get(dimQname)
                            if dimConcept is None:
                                error("xbrl:schemaDefinitionMissing",
                                      _("The taxonomy defined aspect concept QName %(qname)s could not be determined"),
                                      modelObject=modelXbrl, qname=dimQname)
                                continue
                            if dimVal is None:
                                memberAttrs = {"{http://www.w3.org/2001/XMLSchema-instance}nil": "true"}
                            else:
                                memberAttrs = None
                            if isinstance(dimVal, dict):
                                dimVal = dimVal["value"]
                            else:
                                dimVal = str(dimVal) # may be int or boolean
                            if isinstance(dimVal,str) and ":" in dimVal and dimVal.partition(':')[0] in aliases:
                                mem = qname(dimVal, aliases) # explicit dim
                            elif dimConcept.isTypedDimension:
                                # a modelObject xml element is needed for all of the instance functions to manage the typed dim
                                mem = addChild(modelXbrl.modelDocument, dimConcept.typedDomainElement.qname, text=dimVal, attributes=memberAttrs, appendChild=False)
                            else:
                                mem = None # absent typed dimension
                            if mem is not None:
                                qnameDims[dimQname] = DimValuePrototype(modelXbrl, None, dimQname, mem, "segment")
                    _cntx = modelXbrl.createContext(
                                            entityAsQn.namespaceURI,
                                            entityAsQn.localName,
                                            "forever" if period == "forever" else concept.periodType,
                                            None if concept.periodType == "instant" or period == "forever" 
                                                else dateTime(period.rpartition('/')[0], type=DATETIME),
                                            None if period == "forever"
                                                else dateTime(period.rpartition('/')[2], type=DATETIME),
                                            None, # no dimensional validity checking (like formula does)
                                            qnameDims, [], [],
                                            id=cntxId)
                    cntxTbl[cntxKey] = _cntx
                    if firstCntxUnitFactElt is None:
                        firstCntxUnitFactElt = _cntx
                if concept.isNumeric:
                    if "unit" in dimensions:
                        unitKey = dimensions["unit"]
                    else:
                        unitKey = "xbrli:pure" # default unit
                        if "xbrli" not in namespaces:
                            namespaces["xbrli"] = XbrlConst.xbrli
                    if unitKey in unitTbl:
                        _unit = unitTbl[unitKey]
                    else:
                        _unit = None
                        # validate unit
                        unitKeySub = PrefixedQName.sub(UnitPrefixedQNameSubstitutionChar, unitKey)
                        if not UnitPattern.match(unitKeySub):
                            error("{}:invalidUnitString".format(errPrefix),
                                            _("Unit string representation is lexically invalid, %(unit)s"),
                                            modelObject=modelXbrl, unit=unitKey)
                        else:
                            _mul, _sep, _div = unitKey.partition('/')
                            if _mul.startswith('('):
                                _mul = _mul[1:-1]
                            _muls = [u for u in _mul.split('*') if u]
                            if _div.startswith('('):
                                _div = _div[1:-1]
                            _divs = [u for u in _div.split('*') if u]
                            if _muls != sorted(_muls) or _divs != sorted(_divs):
                                error("{}:invalidUnitString".format(errPrefix),
                                                _("Unit string representation measures are not in alphabetical order, %(unit)s"),
                                                modelObject=modelXbrl, unit=unitKey)
                            try:
                                mulQns = [qname(u, namespaces, prefixException=OIMException("oime:unitPrefix",
                                                                                          _("Unit prefix is not declared: %(unit)s"),
                                                                                          unit=u)) 
                                          for u in _muls]
                                divQns = [qname(u, namespaces, prefixException=OIMException("oime:unitPrefix",
                                                                                          _("Unit prefix is not declared: %(unit)s"),
                                                                                          unit=u))
                                          for u in _divs]
                                unitId = 'u-{:02}'.format(len(unitTbl) + 1)
                                for _measures in mulQns, divQns:
                                    for _measure in _measures:
                                        addQnameValue(modelXbrl.modelDocument, _measure)
                                _unit = modelXbrl.createUnit(mulQns, divQns, id=unitId)
                                if firstCntxUnitFactElt is None:
                                    firstCntxUnitFactElt = _unit
                            except OIMException as ex:
                                error(ex.code, ex.message, modelObject=modelXbrl, **ex.msgArgs)
                        unitTbl[unitKey] = _unit
                else:
                    _unit = None
            
                attrs["contextRef"] = _cntx.id
        
                if fact.get("value") is None:
                    attrs[XbrlConst.qnXsiNil] = "true"
                    text = None
                else:
                    text = fact["value"]
                    
                if concept.isNumeric:
                    if _unit is None:
                        return # skip creating fact because unit was invalid
                    attrs["unitRef"] = _unit.id
                    if text is not None: # no decimals for nil value
                        decimals = fact.get("decimals")
                        attrs["decimals"] = decimals if decimals is not None else "INF"
            else:
                text = None #tuple
                    
            if not id:
                id = syntheticFactFormat.format(factNum)
                factNum += 1
            attrs["id"] = id
            if "id" not in fact: # needed for footnote generation
                fact["id"] = id
                    
            # is value a QName?
            if concept.baseXbrliType == "QNameItemType": # renormalize prefix of instance fact
                text = addQnameValue(modelXbrl.modelDocument, qname(text.strip(), namespaces))
    
            f = modelXbrl.createFact(conceptQn, attributes=attrs, text=text, validate=False)
            if firstCntxUnitFactElt is None:
                firstCntxUnitFactElt = f
            
            xmlValidate(modelXbrl, f)
                    
        currentAction = "creating footnotes"
        footnoteLinks = OrderedDict() # ELR elements
        factLocs = {} # index by (linkrole, factId)
        footnoteNbr = 0
        locNbr = 0
        if isCSV or isXL:
            missingFootnotes = footnoteRefFactIds.keys()  - set(
                                    footnote["footnoteId"] for footnote in footnotes) - set(
                                    footnote["factRef"] for footnote in footnotes if "factRef" in footnote)
            if missingFootnotes:
                error("xbrlce:footnoteNotDefined",
                        _("FootnoteId(s) not defined %(footnoteIds)s."),
                        modelObject=modelXbrl, footnoteIds=", ".join(sorted(missingFootnotes)))
        definedInstanceRoles = set()
        footnoteIdsNotReferenced = set()
        undefinedFootnoteTypes = set()
        undefinedFootnoteGroups = set()
        for factOrFootnote in footnotes:
            if isJSON:
                factFootnotes = []
                for ftType, ftGroups in factOrFootnote.get("links", {}).items():
                    factIDs = (factOrFootnote["id"],)
                    if ftType not in linkTypes:
                        undefinedFootnoteTypes.add(ftType)
                    else:
                        for ftGroup, ftTgtIds in ftGroups.items():
                            if ftGroup not in groups:
                                undefinedFootnoteGroups.add(ftGroup)
                            else:
                                footnote = {"group": groups[ftGroup],
                                            "footnoteType": linkTypes[ftType]}
                                if all(tgtId in xbrlNoteTbl for tgtId in ftTgtIds):
                                    # text footnote
                                    footnote["noteRefs"] = ftTgtIds
                                else:
                                    # fact referencing footnote
                                    footnote["factRef"] = ftTgtIds
                                factFootnotes.append(footnote)
            elif isCSV or isXL: # footnotes contains footnote objects
                factFootnotes = (factOrFootnote,)
                factIDs = tuple(sorted(footnoteRefFactIds[factOrFootnote.get("footnoteId")]))
                if not factIDs:
                    footnoteIdsNotReferenced.add(factOrFootnote.get("footnoteId"))
            for footnote in factFootnotes:
                linkrole = footnote.get("group")
                arcrole = footnote.get("footnoteType")
                if not factIDs or not linkrole or not arcrole or not (
                    footnote.get("factRef") or footnote.get("footnote") is not None or footnote.get("noteRefs") is not None):
                    if not linkrole:
                        warning("oime:footnoteMissingLinkrole",
                                        _("FootnoteId has no linkrole %(footnoteId)s."),
                                        modelObject=modelXbrl, footnoteId=footnote.get("footnoteId"))
                    if not arcrole:
                        warning("oime:footnoteMissingArcrole",
                                        _("FootnoteId has no arcrole %(footnoteId)s."),
                                        modelObject=modelXbrl, footnoteId=footnote.get("footnoteId"))
                    continue
                for refType, refValue, roleTypes in (("role", linkrole, modelXbrl.roleTypes),
                                                     ("arcrole", arcrole, modelXbrl.arcroleTypes)):
                    if (not XbrlConst.isStandardRole(refValue) or XbrlConst.isStandardArcrole(refValue)
                        ) and refValue in roleTypes and refValue not in definedInstanceRoles:
                        definedInstanceRoles.add(refValue)
                        hrefElt = roleTypes[refValue][0]
                        href = hrefElt.modelDocument.uri + "#" + hrefElt.id
                        elt = addChild(modelXbrl.modelDocument.xmlRootElement, 
                                       qname(link, refType+"Ref"), 
                                       attributes=(("{http://www.w3.org/1999/xlink}href", href),
                                                   ("{http://www.w3.org/1999/xlink}type", "simple")),
                                       beforeSibling=firstCntxUnitFactElt)
                        href = modelXbrl.modelDocument.discoverHref(elt)
                        if href:
                            _elt, hrefDoc, hrefId = href
                            _defElt = hrefDoc.idObjects.get(hrefId)
                            if _defElt is not None:
                                _uriAttrName = refType + "URI"
                                _uriAttrValue = _defElt.get(_uriAttrName)
                                if _uriAttrValue:
                                    elt.set(_uriAttrName, _uriAttrValue)
                if linkrole not in footnoteLinks:
                    footnoteLinks[linkrole] = addChild(modelXbrl.modelDocument.xmlRootElement, 
                                                       XbrlConst.qnLinkFootnoteLink, 
                                                       attributes={"{http://www.w3.org/1999/xlink}type": "extended",
                                                                   "{http://www.w3.org/1999/xlink}role": linkrole})
                footnoteLink = footnoteLinks[linkrole]
                if (linkrole, factIDs) not in factLocs:
                    locNbr += 1
                    locLabel = "l_{:02}".format(locNbr)
                    factLocs[(linkrole, factIDs)] = locLabel
                    for factId in factIDs:
                        addChild(footnoteLink, XbrlConst.qnLinkLoc, 
                                 attributes={XLINKTYPE: "locator",
                                             XLINKHREF: "#" + factId,
                                             XLINKLABEL: locLabel})
                locFromLabel = factLocs[(linkrole, factIDs)]
                if footnote.get("footnote"):
                    footnoteNbr += 1
                    footnoteToLabel = "f_{:02}".format(footnoteNbr)
                    attrs = {XLINKTYPE: "resource",
                             XLINKLABEL: footnoteToLabel}
                    if footnote.get("language"):
                        attrs[XMLLANG] = footnote["language"]
                    tgtElt = addChild(footnoteLink, XbrlConst.qnLinkFootnote, attributes=attrs)
                    srcElt = etree.fromstring("<footnote xmlns=\"http://www.w3.org/1999/xhtml\">{}</footnote>"
                                              .format(footnote["footnote"]), parser=modelXbrl.modelDocument.parser)
                    if srcElt.__len__() > 0: # has html children
                        setXmlns(modelXbrl.modelDocument, "xhtml", "http://www.w3.org/1999/xhtml")
                    copyIxFootnoteHtml(srcElt, tgtElt, withText=True, isContinChainElt=False)
                elif footnote.get("noteRefs"):
                    footnoteNbr += 1
                    footnoteToLabel = "f_{:02}".format(footnoteNbr)
                    for noteId in footnote.get("noteRefs"):
                        xbrlNote = xbrlNoteTbl[noteId]
                        attrs = {XLINKTYPE: "resource",
                                 XLINKLABEL: footnoteToLabel}
                        try:
                            if "dimensions" in xbrlNote:
                                attrs[XMLLANG] = xbrlNote["dimensions"]["language"]
                            elif "aspects" in xbrlNote:
                                attrs[XMLLANG] = xbrlNote["aspects"]["language"]
                        except KeyError:
                            pass
                        tgtElt = addChild(footnoteLink, XbrlConst.qnLinkFootnote, attributes=attrs)
                        srcElt = etree.fromstring("<footnote xmlns=\"http://www.w3.org/1999/xhtml\">{}</footnote>"
                                                  .format(xbrlNote["value"]), parser=modelXbrl.modelDocument.parser)
                        if srcElt.__len__() > 0: # has html children
                            setXmlns(modelXbrl.modelDocument, "xhtml", "http://www.w3.org/1999/xhtml")
                        copyIxFootnoteHtml(srcElt, tgtElt, withText=True, isContinChainElt=False)
                elif footnote.get("factRef"):
                    factRef = footnote.get("factRef")
                    if (isCSV or isXL) and factRef in footnoteRefFactIds:
                        fact2IDs = tuple(sorted(footnoteRefFactIds[factRef]))
                    else:
                        fact2IDs = tuple(sorted(factRef))
                    if (linkrole, fact2IDs) not in factLocs:
                        locNbr += 1
                        locLabel = "l_{:02}".format(locNbr)
                        factLocs[(linkrole, fact2IDs)] = locLabel
                        for factId in fact2IDs:
                            addChild(footnoteLink, XbrlConst.qnLinkLoc, 
                                     attributes={XLINKTYPE: "locator",
                                                 XLINKHREF: "#" + factId,
                                                 XLINKLABEL: locLabel})
                    footnoteToLabel = factLocs[(linkrole, fact2IDs)]
                footnoteArc = addChild(footnoteLink, 
                                       XbrlConst.qnLinkFootnoteArc, 
                                       attributes={XLINKTYPE: "arc",
                                                   XLINKARCROLE: arcrole,
                                                   XLINKFROM: locFromLabel,
                                                   XLINKTO: footnoteToLabel})
        if isCSV and footnoteIdsNotReferenced:
            warning("xbrlce:footnotesNotReferenced",
                    _("FootnoteId(s) not referenced %(footnoteIds)s."),
                    modelObject=modelXbrl, footnoteIds=", ".join(sorted(footnoteIdsNotReferenced)))
        if footnoteLinks:
            modelXbrl.modelDocument.linkbaseDiscover(footnoteLinks.values(), inInstance=True)

        if undefinedFootnoteTypes:
            error("xbrlje:unknownPrefix",
                  _("These footnote types are not defined in footnoteTypes: %(ftTypes)s."),
                  modelObject=modelXbrl, ftTypes=", ".join(sorted(undefinedFootnoteTypes)))
        if undefinedFootnoteGroups:
            error("xbrlje:unknownPrefix",
                  _("These footnote groups are not defined in groups: %(ftGroups)s."),
                  modelObject=modelXbrl, ftGroups=", ".join(sorted(undefinedFootnoteGroups)))
                    
        currentAction = "done loading facts and footnotes"
        
        #cntlr.addToLog("Completed in {0:.2} secs".format(time.time() - startedAt),
        #               messageCode="loadFromExcel:info")
    except NotOIMException as ex:
        _return = ex # not an OIM document
    except Exception as ex:
        _return = ex
        if isinstance(ex, OIMException):
            error(ex.code, ex.message, modelObject=modelXbrl, **ex.msgArgs)
        else:
            error("arelleOIMloader:error",
                    "Error while %(action)s, error %(error)s\ntraceback %(traceback)s",
                    modelObject=modelXbrl, action=currentAction, error=ex,
                    traceback=traceback.format_tb(sys.exc_info()[2]))
    
    if priorCWD:
        os.chdir(priorCWD) # restore prior current working directory            startingErrorCount = len(modelXbrl.errors)
        
    #if startingErrorCount < len(modelXbrl.errors):
    #    # had errors, don't allow ModelDocument.load to continue
    #    return OIMException("arelleOIMloader:unableToLoad", "Unable to load due to reported errors")

    global lastFilePath, lastFilePath
    lastFilePath = None
    lastFilePathIsOIM = False
    return _return

lastFilePath = None
lastFilePathIsOIM = False

def isOimLoadable(modelXbrl, mappedUri, normalizedUri, filepath, **kwargs):
    global lastFilePath, lastFilePathIsOIM
    lastFilePath = filepath
    lastFilePathIsOIM = False
    _ext = os.path.splitext(filepath)[1]
    if _ext in (".csv", ".json", ".xlsx", ".xls"):
        lastFilePathIsOIM = True
    elif isHttpUrl(normalizedUri) and '?' in _ext: # query parameters and not .json, may be JSON anyway
        with io.open(filepath, 'rt', encoding='utf-8') as f:
            _fileStart = f.read(4096)
        if _fileStart and re.match(r"\s*\{\s*\"documentType\":\s*\"http:\\+/\\+/www.xbrl.org\\+/WGWD\\+/YYYY-MM-DD\\+/xbrl-json\"", _fileStart):
            lastFilePathIsOIM = True
    return lastFilePathIsOIM

def oimLoader(modelXbrl, mappedUri, filepath, *args, **kwargs):
    if filepath != lastFilePath or not lastFilePathIsOIM:
        return None # not an OIM file

    cntlr = modelXbrl.modelManager.cntlr
    cntlr.showStatus(_("Loading OIM file: {0}").format(os.path.basename(filepath)))
    doc = loadFromOIM(cntlr, modelXbrl.error, modelXbrl.warning, modelXbrl, filepath, mappedUri)
    if doc is None:
        return None # not an OIM file
    modelXbrl.loadedFromOIM = True
    return doc

def guiXbrlLoaded(cntlr, modelXbrl, attach, *args, **kwargs):
    if cntlr.hasGui and getattr(modelXbrl, "loadedFromOIM", False):
        from arelle import ModelDocument
        from tkinter.filedialog import askdirectory
        instanceFile = cntlr.uiFileDialog("save",
                title=_("arelle - Save XBRL instance document"),
                initialdir=cntlr.config.setdefault("outputInstanceDir","."),
                filetypes=[(_("XBRL file .xbrl"), "*.xbrl"), (_("XBRL file .xml"), "*.xml")],
                defaultextension=".xbrl")
        if not instanceFile:
            return False
        cntlr.config["outputInstanceDir"] = os.path.dirname(instanceFile)
        cntlr.saveConfig()
        if instanceFile:
            modelXbrl.modelDocument.save(instanceFile, updateFileHistory=False)
            cntlr.showStatus(_("Saving XBRL instance: {0}").format(os.path.basename(instanceFile)))
        cntlr.showStatus(_("OIM loading completed"), 3500)

def cmdLineXbrlLoaded(cntlr, options, modelXbrl, *args, **kwargs):
    if options.saveOIMinstance and getattr(modelXbrl, "loadedFromOIM", False):
        doc = modelXbrl.modelDocument
        cntlr.showStatus(_("Saving XBRL instance: {0}").format(doc.basename))
        responseZipStream = kwargs.get("responseZipStream")
        if responseZipStream is not None:
            _zip = zipfile.ZipFile(responseZipStream, "a", zipfile.ZIP_DEFLATED, True)
        else:
            _zip = None
        doc.save(options.saveOIMinstance, _zip)
        if responseZipStream is not None:
            _zip.close()
            responseZipStream.seek(0)

def excelLoaderOptionExtender(parser, *args, **kwargs):
    parser.add_option("--saveOIMinstance", 
                      action="store", 
                      dest="saveOIMinstance", 
                      help=_("Save a instance loaded from OIM into this file name."))
    
def oimJsonSaveXml(cntlr, oimJsonObject, jsonFileName, xbrlFileName):
    def _error(code, message, **kwargs):
        cntlr.addToLog(message, code, kwargs, level=logging.ERROR)
    def _warning(code, message, **kwargs):
        cntlr.addToLog(message, code, kwargs, level=logging.WARNING)
        
    doc = loadFromOIM(cntlr, _error, _warning, None, jsonFileName, "OIM", oimJsonObject)
    if xbrlFileName:
        doc.save(xbrlFileName)
    doc.modelXbrl.close()
    
__pluginInfo__ = {
    'name': 'Load From OIM',
    'version': '1.2',
    'description': "This plug-in loads XBRL instance data from OIM (JSON, CSV or Excel) and saves the resulting XBRL Instance.",
    'license': 'Apache-2',
    'author': 'Mark V Systems Limited',
    'copyright': '(c) Copyright 2016 Mark V Systems Limited, All rights reserved.',
    # classes of mount points (required)
    'ModelDocument.IsPullLoadable': isOimLoadable,
    'ModelDocument.PullLoader': oimLoader,
    'CntlrWinMain.Xbrl.Loaded': guiXbrlLoaded,
    'CntlrCmdLine.Options': excelLoaderOptionExtender,
    'CntlrCmdLine.Xbrl.Loaded': cmdLineXbrlLoaded,
    'OimJson.SaveXml': oimJsonSaveXml
}
